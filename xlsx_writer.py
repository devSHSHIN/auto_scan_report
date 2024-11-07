from openpyxl import Workbook

def write_data_to_sheet(ws, sheet, value):
    ws.title = sheet

    if isinstance(value, dict):
        row = 1
        for s, v in value.items():
            ws.cell(row=row, column=1, value=s)
            ws.cell(row=row, column=1, value=str(v))
            row += 1
    
    elif isinstance(value, list):
        row = 1
        for item in value:
            ws.cell(row=row, column=1, value=str(item))
            row += 1

    else:
        ws.cell(row=1, column=1, value=value)

def save_xlsx_file(workbook, xlsx_file_path):
    workbook.save(xlsx_file_path)
    print(f"엑셀 파일이 저장되었습니다 :\t[ {xlsx_file_path} ]")
