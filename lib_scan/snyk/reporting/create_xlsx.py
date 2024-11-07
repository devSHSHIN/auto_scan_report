import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill

guide_list = [
    "build.gradle(Gradle), pom.xl(maven) 등의 라이브러리/의존성 설정 파일에서 표에 적힌 Package의 버전을 현재버전에서 패치버전으로 수정해 주시길 바랍니다.",
    "해당 라이브러리의 추가적인 정보를 확인하고자 한다면, 위 사이트에서 Packege에 해당하는 내용을 검색하여 조치에 참고하시길 바랍니다.",
    "조치 후 JIRA를 통해 Comment를 남겨주시길 바랍니다.",
    "",
    "※ 라이브러리 취약점은 매년 1회만 정기점검을 실시하여 전달드립니다.",
    "따라서, 1회에 한하여 취약점 공유 및 조치확인을 하게 됩니다. 다만, 취약점 공유 이후에 업데이트필요성이 아주 높은 취약점이 발견되면",
    "해당 년도에 추가적으로 한번 더 업데이트를 요청 드릴 수 있음을 알려드립니다.",
    "(라이브러리의 경우 잦은 업데이트가 일어남에 따라 신규 취약점이 지속적으로 발견되므로, 항상 최신 업데이트를 유지 할수 있도록 추적 관리 해주시길 바랍니다)",
    "",
    "[정규 진단 1차]",
    "",
    "취약한 버전의 패키지를 포함하는 모듈을 사용하고 있음",
    "영향성을 검토하여 취약점이 조치된 버전으로 패키지 업데이트 권고드립니다.",
    ""
]

title = [
    "",
    "No",
    "플랫폼구분",
    "CWE Name",
    "점검결과",
    "진단일자",
    "이행점검결과",
    "Severity Level",
    "Package",
    "현재버전",
    "패치버전",
    "비고",
    "조치계획",
    "조치내역",
    "조치일자",
    "조치 Commit No"
]

def load_json(issue_path):
    return pd.read_json(issue_path)

def load_txt(log_path):
    with open(log_path, 'r') as f:
        lines = f.readlines()
    
    log_content = ''.join(lines)

    line_cnt = log_content.count('\n') + 1

    return log_content, line_cnt

def modify_data(data):
    if data.empty:
        print("발견된 취약점이 없습니다.")
        return None  # 또는 시스템 종료 시 `sys.exit()` 사용 가능

    data.iloc[:, 4] = "취약"
    data.iloc[:, 11] = data.iloc[:, 11].apply(lambda x: f"외 {int(x)}개" if pd.notna(x) and isinstance(x, (int, float)) and x > 0 else "")

    data.iloc[0, 6] = "-"
    data.iloc[0, 12] = "ex) yyyy-mm-dd"
    data.iloc[0, 13] = "ex) 상세 조치 내용 기재"
    data.iloc[0, 14] = "ex) 조치완료/조치예정"

    return data

def set_title(ws):
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )
    center = Alignment(horizontal="center", vertical="center")
    font = Font(bold=True, name='맑은 고딕', size=11, color="000000")

    ws.append(title)

    for row in ws['B2:P2']:
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = center
            cell.font = font
            
def set_table(ws, data):
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    center = Alignment(horizontal="center", vertical="center")
    font = Font(name='맑은 고딕', size=10, color="000000")

    for issue in data.itertuples(index=False, name=None):
        issue = [str(item) if isinstance(item, list) else item for item in issue]
        ws.append(issue)
        now_row = f"B{ws.max_row}:P{ws.max_row}"

        for row in ws[now_row]:
            for cell in row:
                cell.border = border
                cell.font = font
                cell.alignment = center

                if "D" in cell.coordinate or "I" in cell.coordinate:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif "E" in cell.coordinate:
                    cell.font = Font(bold=True, name='맑은 고딕', size=10, color="FF0000")
                elif cell.coordinate[0] in ("M", "N", "O", "P"):
                    cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                    cell.font = Font(bold=True, name='맑은 고딕', size=10, color="FF0000")


def add_row(ws, data=""):
    ws.append(["", data])
    ws[f"B{ws.max_row}"].font = Font(bold=True, name='맑은 고딕', size=10)
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=16)

    return ws.max_row

def create_xlsx(issue_path, log_path):
    issue_tmp = load_json(issue_path)
    issue_list = modify_data(issue_tmp)

    log_txt, log_cnt = load_txt(log_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Snyk Issues"

    add_row(ws)

    set_title(ws)
    set_table(ws, issue_list)

    merge_cells = f"B2:P{ws.max_row}"

    add_row(ws)
    add_row(ws)
    black_row = add_row(ws,"점검 내용 상세")
    add_row(ws)
    
    ws.append(["", "조치 방법"])

    for txt in guide_list:
        add_row(ws, txt)

    log_row = ws.max_row + 1
    log_column = 2

    ########    스타일 서식 적용    ########

    ws.cell(row=black_row, column=2).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.cell(row=black_row, column=2).font = Font(bold=True, name='맑은 고딕', size=11, color="FFFFFF")

    url_row = black_row + 2
    url_cell = f"D{url_row}"

    ws[f"B{url_row}"].font = Font(bold=True, name="맑은 고딕", size=11, color="FF0000")
    ws[url_cell].value = "MVN Repository"
    ws[url_cell].hyperlink = "https://mvnrepository.com/"
    ws[url_cell].style = "Hyperlink"
    ws[f"B{url_row + 10}"].font = Font(bold=True, name="맑은 고딕", size=11, color="000000")

    for col in range(2, 16):
        ws.cell(row=url_row, column=col).border = Border(top=Side(border_style="thin", color="000000"))
    for col in range(2, 16):
        ws.cell(row=url_row + 3, column=col).border = Border(bottom=Side(border_style="thin", color="000000"))
    for r in range(url_row + 1, url_row + 3):
        ws.cell(row=r, column=2).border = Border(left=Side(border_style="thin", color="000000"))
    for r in range(url_row + 1, url_row + 3):
        ws.cell(row=r, column=16).border = Border(right=Side(border_style="thin", color="000000"))
    
    ws.merge_cells(start_row=url_row, start_column=2, end_row=url_row, end_column=3)
    ws.merge_cells(start_row=url_row, start_column=4, end_row=url_row, end_column=16)

    ws.merge_cells(start_row=log_row, start_column=log_column, end_row=log_row + log_cnt + 1, end_column=16)
    ws.cell(row=log_row, column=log_column, value=log_txt)
    ws.cell(row=log_row, column=log_column).alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells(start_row=2, start_column=1, end_row=ws.max_row, end_column=1)
    ws.column_dimensions['A'].width = 3

    # 엑셀 파일 저장 (경로 수정)
    output_path = "/Users/pc09164/auto_scan_report/data/Snyk_Report.xlsx"
    wb.save(output_path)
    wb.close()

    return output_path
