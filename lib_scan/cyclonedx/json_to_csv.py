import json
from openpyxl import Workbook

# JSON 데이터를 엑셀로 변환하는 함수
def json_to_excel(json_data, output_file):
    wb = Workbook()
    ws = wb.active

    # JSON 데이터를 재귀적으로 처리하는 함수
    def process_node(data, depth=3, row=1):
        if isinstance(data, dict):
            for key, value in data.items():
                # C열부터 계층적으로 데이터 입력
                ws.cell(row=row, column=depth, value=key)
                row = process_node(value, depth + 1, row + 1)
        elif isinstance(data, list):
            for item in data:
                row = process_node(item, depth, row)
        else:
            # 기본 데이터(값)를 처리하여 셀에 입력
            ws.cell(row=row, column=depth, value=data)
            row += 1
        return row

    # JSON 데이터 처리 시작
    process_node(json_data)

    # 엑셀 파일 저장
    wb.save(output_file)
    print(f"엑셀 파일이 생성되었습니다: {output_file}")

# JSON 파일을 엑셀로 변환하는 함수 실행 예시
def convert_json_file_to_excel(input_json_file, output_excel_file):
    # JSON 파일 읽기
    with open(input_json_file, 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    # JSON 데이터를 엑셀 파일로 변환
    json_to_excel(json_data, output_excel_file)

# 사용 예시
#input_file = 'input.json'   # JSON 파일 경로
#output_file = 'output.xlsx' # 출력할 엑셀 파일 경로
input_file = input(f"Input_JSON :\t")
output_file = input(f"Output_xlsx :\t")

convert_json_file_to_excel(input_file, output_file)
